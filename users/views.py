from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import UserProfile, Appliance, EnergyConsumption, Prediction, Report
from django.utils import timezone
from .services import generate_dummy_prediction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.utils import get_column_letter
import json

# API for frontend registration via fetch()
@csrf_exempt
def register_user(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get('name')
        contact = data.get('contact')
        email = data.get('email')
        address = data.get('address')
        password = data.get('password')

        print("Received registration:", name, contact, email)

        # Check for duplicates
        if User.objects.filter(email=email).exists():
            return JsonResponse({'status': 'error', 'message': 'Email already exists'})
        if UserProfile.objects.filter(contact_number=contact).exists():
            return JsonResponse({'status': 'error', 'message': 'Contact number already exists'})

        # Create user and profile
        user = User.objects.create_user(username=email, email=email, password=password)
        user.save()
        user_profile = UserProfile(user=user, full_name=name, contact_number=contact, physical_address=address)
        user_profile.save()

        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Only POST allowed'})

# Export user data to Excel
def export_user_data_to_excel(request):
    if not request.user.is_staff:
        return HttpResponse("Unauthorized", status=401)

    user_profiles = UserProfile.objects.prefetch_related('appliances').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Profiles"

    headers = ['Full Name', 'Contact Number', 'Physical Address', 'Email', 'Appliances']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    for row_num, profile in enumerate(user_profiles, 2):
        appliances_str = ", ".join([f"{a.name} (Qty: {a.quantity}, Power: {a.power_rating}W)" for a in profile.appliances.all()])
        ws[f'A{row_num}'] = profile.full_name
        ws[f'B{row_num}'] = profile.contact_number
        ws[f'C{row_num}'] = profile.physical_address
        ws[f'D{row_num}'] = profile.user.email
        ws[f'E{row_num}'] = appliances_str

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=user_profiles.xlsx'
    wb.save(response)
    return response

# Signup view for users
def signup_view(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        contact_number = request.POST.get('contact_number')
        email = request.POST.get('email')
        physical_address = request.POST.get('physical_address')
        password = request.POST.get('password')
        password2 = request.POST.get('password2')
        # captcha_input = request.POST.get('captchaInput')

        # Temporarily disable CAPTCHA validation for testing
        # if not captcha_input or captcha_input.lower() != 'expectedcaptcha':
        #     messages.error(request, "Invalid CAPTCHA. Please try again.")
        #     return render(request, 'signup.html')

        if password != password2:
            messages.error(request, "Passwords do not match")
            return render(request, 'signup.html')

        user = None
        user_profile = None

        # Check if user with email exists
        try:
            user = User.objects.get(email=email)
            user_profile = UserProfile.objects.get(user=user)
            # Check if contact_number is used by another user_profile
            if UserProfile.objects.filter(contact_number=contact_number).exclude(user=user).exists():
                messages.error(request, "Contact number already exists for another user")
                return render(request, 'signup.html')

            # Update existing user and profile
            user.username = email
            user.set_password(password)
            user.email = email
            user.save()

            user_profile.full_name = full_name
            user_profile.contact_number = contact_number
            user_profile.physical_address = physical_address
            user_profile.save()

            messages.success(request, "Account updated successfully")
            return redirect('add_appliances')
        except User.DoesNotExist:
            # Create new user and profile
            if UserProfile.objects.filter(contact_number=contact_number).exists():
                messages.error(request, "Contact number already exists")
                return render(request, 'signup.html')

            user = User.objects.create_user(username=email, email=email, password=password)
            user.save()
            user_profile = UserProfile(user=user, full_name=full_name, contact_number=contact_number, physical_address=physical_address)
            user_profile.save()
            # Authenticate and login the new user
            authenticated_user = authenticate(request, username=email, password=password)
            if authenticated_user is not None:
                login(request, authenticated_user)
            messages.success(request, "Account created successfully")
            return redirect('add_appliances')

    return render(request, 'signup.html')

# Login view for users
def user_login(request):
    if request.method == 'POST':
        contact_no = request.POST.get('contact_no')
        password = request.POST.get('password')

        # Check for admin login
        if contact_no == '9874563210' and password == 'Aashi':
            try:
                user = User.objects.get(username='admin@example.com')  # Use actual admin username/email
                login(request, user)
                return redirect('admin_dashboard')
            except User.DoesNotExist:
                messages.error(request, 'Admin account not found. Please create it in the admin panel.')

        # Check for regular user login
        try:
            user_profile = UserProfile.objects.get(contact_number=contact_no)
            # Authenticate using email as username
            user = authenticate(request, username=user_profile.user.email, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, 'Invalid contact number or password.')
        except UserProfile.DoesNotExist:
            messages.error(request, 'Invalid contact number or password.')

    return render(request, 'login.html')

# Admin dashboard view (no staff data)
@login_required(login_url='login')
def admin_dashboard_view(request):
    user_profiles = UserProfile.objects.prefetch_related('appliances').all()
    return render(request, 'admin_dashboard.html', {'user_profiles': user_profiles})

# User dashboard view
from django.db.models import Sum, F, ExpressionWrapper, FloatField, DateField
from django.db.models.functions import TruncDay
from datetime import timedelta, datetime

from django.views.decorators.cache import never_cache

@never_cache
@login_required(login_url='login')
def dashboard_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    appliances = Appliance.objects.filter(user_profile=user_profile)
    predictions = Prediction.objects.filter(user_profile=user_profile).order_by('-predicted_date')[:5]
    reports = Report.objects.filter(user_profile=user_profile).order_by('-report_date')[:5]

    # Calculate total consumption, total cost, and efficiency (dummy calculations for now)
    total_consumption = 0
    total_cost = 0
    efficiency = 0

    # Calculate consumption per appliance
    consumption_per_appliance = []
    for appliance in appliances:
        usage_time = getattr(appliance, 'usage_time', 1)  # default 1 hour if not available
        days_used_per_week = getattr(appliance, 'days_used_per_week', 7)  # default 7 days if not available
        # Calculate monthly consumption: power_rating * quantity * usage_time * (days_used_per_week * 4.33 / 7)
        consumption = appliance.power_rating * appliance.quantity * usage_time * (days_used_per_week * 4.33 / 7)
        total_consumption += consumption
        consumption_per_appliance.append(round(consumption, 2))

    # Dummy cost calculation: total_consumption * rate_per_kwh (assume 0.12)
    rate_per_kwh = 0.12
    total_cost = total_consumption * rate_per_kwh

    # Dummy efficiency calculation: arbitrary formula or placeholder
    if total_consumption > 0:
        efficiency = (total_consumption / (total_consumption + 100)) * 100
    else:
        efficiency = 0

    # Serialize appliances for JSON script in template
    appliances_data = []
    for idx, appliance in enumerate(appliances):
        appliances_data.append({
            'id': appliance.id,
            'name': appliance.name,
            'power_rating': appliance.power_rating,
            'quantity': appliance.quantity,
            'consumption': consumption_per_appliance[idx],
            'cost': round(consumption_per_appliance[idx] * rate_per_kwh, 2),
        })

    # Aggregate appliance-wise total consumption (kWh) and cost
    appliance_consumption_totals = appliances.annotate(
        total_consumption=ExpressionWrapper(
            F('power_rating') * F('quantity') * 1,  # Assuming 1 hour usage for now
            output_field=FloatField()
        )
    ).values('name').annotate(total=Sum('total_consumption')).order_by('name')

    # Appliance-wise consumption and cost for current user
    appliance_names = [appliance['name'] for appliance in appliances_data]
    appliance_consumptions = [appliance['consumption'] for appliance in appliances_data]
    appliance_costs = [appliance['cost'] for appliance in appliances_data]

    # Dummy dynamic data for prediction section
    predicted_total_usage = round(total_consumption * 1.1, 2)  # example: 10% increase forecast
    day1_usage = round(total_consumption * 0.15, 2)
    day1_temp = 38
    day1_cost = round(day1_usage * rate_per_kwh, 2)
    day2_usage = round(total_consumption * 0.12, 2)
    day2_temp = 32
    day2_cost = round(day2_usage * rate_per_kwh, 2)
    day3_usage = round(total_consumption * 0.10, 2)
    day3_temp = 30
    day3_cost = round(day3_usage * rate_per_kwh, 2)
    ai_savings = 600  # example static value

    context = {
        'full_name': user_profile.full_name,
        'contact_number': user_profile.contact_number,
        'email': request.user.email,
        'physical_address': user_profile.physical_address,
        'appliances': appliances,
        'appliances_data': appliances_data,
        'appliance_names': appliance_names,
        'appliance_power_ratings': [appliance['power_rating'] for appliance in appliances_data],
        'appliance_quantities': [appliance['quantity'] for appliance in appliances_data],
        'appliance_consumptions': appliance_consumptions,
        'appliance_costs': appliance_costs,
        'predictions': predictions,
        'reports': reports,
        'total_consumption': round(total_consumption, 2),
        'total_cost': round(total_cost, 2),
        'efficiency': round(efficiency, 2),
        'appliance_consumption_totals': list(appliance_consumption_totals),
        'predicted_total_usage': predicted_total_usage,
        'day1_usage': day1_usage,
        'day1_temp': day1_temp,
        'day1_cost': day1_cost,
        'day2_usage': day2_usage,
        'day2_temp': day2_temp,
        'day2_cost': day2_cost,
        'day3_usage': day3_usage,
        'day3_temp': day3_temp,
        'day3_cost': day3_cost,
        'ai_savings': ai_savings,
    }
    return render(request, 'dashboard.html', context)

# Add appliances view for users
@login_required(login_url='login')
def add_appliances_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    if request.method == 'POST':
        names = request.POST.getlist('name[]')
        brands = request.POST.getlist('brand[]')
        models = request.POST.getlist('model[]')
        usage_times = request.POST.getlist('usage_time[]')
        quantities = request.POST.getlist('quantity[]')
        days_used_per_weeks = request.POST.getlist('days_used_per_week[]')
        power_ratings = request.POST.getlist('power_rating[]')
        if not (names and brands and models and usage_times and quantities and days_used_per_weeks and power_ratings):
            messages.error(request, "All fields are required.")
            return render(request, 'add_appliances.html')

        # Instead of deleting existing appliances, add new ones cumulatively
        for i in range(len(names)):
            try:
                name = names[i]
                brand = brands[i]
                model = models[i]
                usage_time = float(usage_times[i])
                quantity = int(quantities[i])
                days_used_per_week = int(days_used_per_weeks[i])
                power_rating = float(power_ratings[i])

                # Check if appliance with same name exists for user
                appliance_name = f"{name} - {brand} {model}"
                existing_appliance = Appliance.objects.filter(user_profile=user_profile, name=appliance_name).first()
                if existing_appliance:
                    # Increment quantity if appliance exists
                    existing_appliance.quantity += quantity
                    existing_appliance.days_used_per_week = days_used_per_week
                    existing_appliance.usage_time = usage_time
                    existing_appliance.power_rating = power_rating
                    existing_appliance.save()
                else:
                    # Create new appliance
                    Appliance.objects.create(
                        user_profile=user_profile,
                        name=appliance_name,
                        power_rating=power_rating,
                        quantity=quantity,
                        days_used_per_week=days_used_per_week,
                        usage_time=usage_time
                    )
            except (ValueError, IndexError):
                messages.error(request, "Invalid appliance data provided.")
                return render(request, 'add_appliances.html')

        messages.success(request, "Appliances added successfully.")
        return redirect('dashboard')

    return render(request, 'add_appliances.html')

# Logout view
def logout_view(request):
    logout(request)
    return redirect('login')

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponseForbidden

@login_required(login_url='login')
def remove_appliance_view(request, appliance_id):
    if request.method == 'POST':
        appliance = get_object_or_404(Appliance, id=appliance_id)
        user_profile = Appliance.objects.get(id=appliance_id).user_profile
        if user_profile.user == request.user:
            appliance.delete()
            return redirect('dashboard')
        else:
            return HttpResponseForbidden("You are not allowed to delete this appliance.")
    else:
        return HttpResponseForbidden("Invalid request method.")

# Admin remove appliance for any user
from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def admin_remove_appliance_view(request, appliance_id):
    if request.method == 'POST':
        appliance = get_object_or_404(Appliance, id=appliance_id)
        appliance.delete()
        return redirect('admin_dashboard')
    else:
        return HttpResponseForbidden("Invalid request method.")

# Admin add appliance for any user
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponseForbidden

@staff_member_required
def admin_add_appliance_view(request):
    if request.method == 'POST':
        user_profile_id = request.POST.get('user_profile_id')
        name = request.POST.get('name')
        power_rating = request.POST.get('power_rating')
        quantity = request.POST.get('quantity')

        if not (user_profile_id and name and power_rating and quantity):
            messages.error(request, "All fields are required.")
            return redirect('admin_dashboard')

        try:
            user_profile = UserProfile.objects.get(id=user_profile_id)
            power_rating = float(power_rating)
            quantity = int(quantity)
        except (UserProfile.DoesNotExist, ValueError):
            messages.error(request, "Invalid data provided.")
            return redirect('admin_dashboard')

        # Check if appliance with same name exists for user
        existing_appliance = Appliance.objects.filter(user_profile=user_profile, name=name).first()
        if existing_appliance:
            existing_appliance.quantity += quantity
            existing_appliance.power_rating = power_rating  # Update power rating if changed
            existing_appliance.save()
        else:
            Appliance.objects.create(
                user_profile=user_profile,
                name=name,
                power_rating=power_rating,
                quantity=quantity
            )
        messages.success(request, "Appliance added successfully.")
        return redirect('admin_dashboard')
    else:
        return HttpResponseForbidden("Invalid request method.")

# Admin remove user and all associated data
@staff_member_required
def admin_remove_user_view(request, user_profile_id):
    if request.method == 'POST':
        user_profile = get_object_or_404(UserProfile, id=user_profile_id)
        user = user_profile.user
        # Delete user profile and user (cascade should delete related data)
        user.delete()
        messages.success(request, f"User {user_profile.full_name} and all associated data removed successfully.")
        return redirect('admin_dashboard')
    else:
        return HttpResponseForbidden("Invalid request method.")

# View to list appliances
@login_required(login_url='login')
def appliances_list_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    appliances = Appliance.objects.filter(user_profile=user_profile)
    context = {
        'appliances': appliances,
    }
    return render(request, 'appliances_list.html', context)

# View for consumption of an appliance
@login_required(login_url='login')
def consumption_view(request, appliance_id):
    try:
        appliance = Appliance.objects.get(id=appliance_id)
    except Appliance.DoesNotExist:
        messages.error(request, "Appliance not found")
        return redirect('dashboard')
    consumptions = EnergyConsumption.objects.filter(appliance=appliance).order_by('-date')
    context = {
        'appliance': appliance,
        'consumptions': consumptions,
    }
    return render(request, 'consumption.html', context)

# View for predictions
@login_required(login_url='login')
def predictions_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    predictions = Prediction.objects.filter(user_profile=user_profile).order_by('-predicted_date')
    context = {
        'predictions': predictions,
    }
    return render(request, 'predictions.html', context)

# View for reports
@login_required(login_url='login')
def reports_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    reports = Report.objects.filter(user_profile=user_profile).order_by('-report_date')
    context = {
        'reports': reports,
    }
    return render(request, 'reports.html', context)

# Generate predictions view
@login_required(login_url='login')
def generate_predictions_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    generate_dummy_prediction(user_profile)
    messages.success(request, "Predictions generated successfully")
    return redirect('predictions')
